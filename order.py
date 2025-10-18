import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime
import pandas as pd


# --- Настройки ---
BASE_URL = "https://www.beelifecos.com/beelifecos_admin.php"
ORDER_URL = BASE_URL + "?dispatch=orders.print_invoice&order_id={}"
USERNAME = "beelifecos@gmail.com"
PASSWORD = "beelifecos140512"
ORDER_IDS = [3845]  # список заказов

desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
orders_path = os.path.join(desktop_path, "orders")
os.makedirs(orders_path, exist_ok=True)
file_name = f"invoce_packing_{'_'.join(map(str, ORDER_IDS))}.xlsx"
file_path = os.path.join(orders_path, file_name)

# --- Сессия и авторизация ---
session = requests.Session()
headers = {
    'User-Agent': 'Mozilla/5.0',
    'Accept': 'text/html',
}

resp = session.get(BASE_URL, headers=headers)
soup = BeautifulSoup(resp.text, "html.parser")
form = soup.find("form", {"name": "main_login_form"})
login_data = {inp.get("name"): inp.get("value", "") for inp in form.find_all("input") if inp.get("name")}
login_data["user_login"] = USERNAME
login_data["password"] = PASSWORD
login_data["dispatch[auth.login]"] = "Войти"
session.post(BASE_URL, data=login_data, headers=headers)

# --- Функция для получения данных заказа ---
def fetch_order(order_id):
    response = session.get(ORDER_URL.format(order_id), headers=headers)
    soup = BeautifulSoup(response.text, "html.parser")
    table = soup.find("table", width="100%")
    
    # Значения по умолчанию
    customer_name = "UNKNOWN"
    customer_address = "UNKNOWN"
    customer_country = "UNKNOWN"
    text = ""  

    # Берём текст из blockquote
    blockquote = soup.find("blockquote")
    if blockquote:
        text = blockquote.get_text(" ", strip=True)
    
    # Имя
    if "Здравствуйте," in text:
        customer_name = text.split("Здравствуйте,")[-1].split("адрес:")[0].replace("!", "").strip()
    
    # Адрес
    if "адрес:" in text:
        customer_address = text.split("адрес:")[-1].split("страна:")[0].strip()
    
    # Страна
    if "страна:" in text:
        customer_country = text.split("страна:")[-1].strip()

    # Трек номер
    track_number = "UNKNOWN"
    td_tags = soup.find_all("td", style="vertical-align: top;")
    for td in td_tags:
        text = td.get_text(" ", strip=True)
        if "Трек номер" in text:
            track_number = text.split("Трек номер:")[-1].strip()
            break
   
    # Товары
    rows_data = []
    if table:
        for tr in table.find_all("tr")[1:]:
            tds = tr.find_all("td")
            if len(tds) < 7:
                continue
            p_tag = tds[2].find("p")
            name = p_tag.get_text(strip=True) if p_tag else tds[2].get_text(" ", strip=True)
            article = tds[1].get_text(strip=True)
            img_tag = tds[0].find("img")
            link = img_tag["src"] if img_tag else ""
            try:
                qty_collected = int(tds[3].get_text(strip=True))
            except:
                qty_collected = 0
            try:
                qty_order = int(tds[4].get_text(strip=True))
            except:
                qty_order = 0
            try:
                price_text = tds[5].get_text(strip=True).replace("$","").replace(",",".")
                price = float(price_text)
            except:
                price = 0.0
            discount_text = tds[6].get_text(strip=True)
            discount = float(discount_text.replace("$","").replace(",", ".")) if discount_text != "-" else 0.0
            rows_data.append([article, name, qty_collected, qty_order, price, discount, link])
    return customer_name, track_number, rows_data, customer_address, customer_country

# --- Объединяем все заказы ---
client_data = {}
for oid in ORDER_IDS:
    customer_name, track_number, rows, address, country = fetch_order(oid)
    key = customer_name
    if key not in client_data:
        client_data[key] = {"track": track_number, "rows": [], "order_ids": [], "address": address, "country": country}
    client_data[key]["rows"].extend(rows)
    client_data[key]["order_ids"].append(str(oid))

# --- Создаём Excel ---
wb = Workbook()
bold_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for client, data in client_data.items():
    # --- Лист 1: Все заказы ---
    ws_orders = wb.active
    ws_orders.title = "Все заказы"
    ws_orders["A1"], ws_orders["B1"] = "Номер заказа:", ", ".join(data["order_ids"])
    ws_orders["A2"], ws_orders["B2"] = "Имя заказчика:", client
    ws_orders["A3"], ws_orders["B3"] = "Трек номер:", data["track"]
    ws_orders["C1"], ws_orders["D1"] = "Адрес:", data["address"]
    ws_orders["C2"], ws_orders["D2"] = "Страна:", data["country"]

    for i in range(1,4):
        ws_orders[f"A{i}"].font = bold_font
        ws_orders[f"A{i}"].alignment = Alignment(horizontal="left", vertical="center")
    
    excel_headers = ["Фото", "Артикул", "Название", "Кол-во собрано", "Кол-во заказа", "Цена", "Скидка", "Сумма заказа", "Сумма отгруженного"]
    ws_orders.append(excel_headers)
    for col in range(1, len(excel_headers)+1):
        cell = ws_orders.cell(row=4, column=col)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    start_row = 5
    for idx, row in enumerate(data["rows"], start=start_row):
        article, name, qty_collected, qty_order, price, discount, img_url = row
        ws_orders.append(["", article, name, qty_collected, qty_order, price, discount, "", ""])
        ws_orders[f"H{idx}"] = f"=(F{idx}-G{idx})*E{idx}"
        ws_orders[f"I{idx}"] = f"=(F{idx}-G{idx})*D{idx}"
        fill = red_fill if qty_collected==0 else yellow_fill if qty_collected<qty_order else green_fill
        for col in range(1,10):
            ws_orders.cell(row=idx, column=col).fill = fill
            ws_orders.cell(row=idx, column=col).border = thin_border
            ws_orders.cell(row=idx, column=col).alignment = center_align if col!=3 else Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws_orders.row_dimensions[idx].height = 50
        if img_url:
            try:
                img_data = requests.get(img_url).content
                image = Image(BytesIO(img_data))
                scale = min(13*7/image.width, 50/image.height)
                image.width = int(image.width*scale)
                image.height = int(image.height*scale)
                ws_orders.add_image(image, f"A{idx}")
            except: pass

    last_data_row = start_row + len(data["rows"]) -1
    ws_orders[f"H{last_data_row+1}"] = "Итого заказ"
    ws_orders[f"I{last_data_row+1}"] = f"=SUM(H{start_row}:H{last_data_row})"
    ws_orders[f"H{last_data_row+2}"] = "Итого отгружено"
    ws_orders[f"I{last_data_row+2}"] = f"=SUM(I{start_row}:I{last_data_row})"
    ws_orders[f"H{last_data_row+3}"] = "Доставка"
    ws_orders[f"I{last_data_row+3}"] = 0
    ws_orders[f"H{last_data_row+4}"] = "Скидка"
    ws_orders[f"I{last_data_row+4}"] = 0
    ws_orders[f"H{last_data_row+5}"] = "Итого к оплате"
    ws_orders[f"I{last_data_row+5}"] = f"=I{last_data_row+2}+I{last_data_row+3}-I{last_data_row+4}"
    ws_orders.column_dimensions["C"].width = 50
    for col_letter in ["D","E","F","G","H","I"]:
        ws_orders.column_dimensions[col_letter].width = 13

# --- Лист 2: Packing List ---
ws_pl = wb.create_sheet("Packing List")
ws_pl["A1"], ws_pl["B1"] = "Имя заказчика:", client
ws_pl["A2"], ws_pl["B2"] = "Номер заказа:", ", ".join(data["order_ids"])
ws_pl["A3"], ws_pl["B3"] = "Трек номер:", data["track"]
for i in range(1,4):
    ws_pl[f"A{i}"].font = bold_font
    ws_pl[f"A{i}"].alignment = Alignment(horizontal="left", vertical="center")

# Заголовки с нужными колонками
pl_headers = ["Бренд", "Наименование товара", "Баркод", "Вес", "Изготовитель", "Адрес завода изготовителя", "Кол-во отгружено"]
ws_pl.append(pl_headers)
for col in range(1, len(pl_headers)+1):
    cell = ws_pl.cell(row=4, column=col)
    cell.font = white_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

pl_start = 5

for row in data["rows"]:
    barcode = row[0]
    name = row[1]
    qty_collected = row[2]

    weight = "_"
    manufacturer = "-"
    factory_address = "-"

    if qty_collected > 0:
        # Формула ВПР для бренда
        vlookup_formula = f'=VLOOKUP(C{pl_start};\'Purchase Invoice\'!B:M;12;FALSE)'
        ws_pl.append([vlookup_formula, name, barcode, weight, manufacturer, factory_address, qty_collected])

        ws_pl.row_dimensions[pl_start].height = 50
        for col_idx in range(1, len(pl_headers)+1):
            ws_pl.cell(row=pl_start, column=col_idx).alignment = center_align if col_idx != 2 else Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws_pl.cell(row=pl_start, column=col_idx).border = thin_border

        pl_start += 1
        
ws_pl[f"A{pl_start}"] = "Общее кол-во отгруженного"
ws_pl[f"G{pl_start}"] = f"=SUM(G5:G{pl_start-1})"
ws_pl[f"A{pl_start}"].font = bold_font
ws_pl[f"G{pl_start}"].font = bold_font
ws_pl[f"A{pl_start}"].alignment = Alignment(horizontal="left", vertical="center")
ws_pl[f"G{pl_start}"].alignment = center_align

# Подгоняем ширину колонок
column_widths = [20, 50, 20, 10, 25, 35, 15]
for i, width in enumerate(column_widths, start=1):
    ws_pl.column_dimensions[chr(64+i)].width = width


    # --- Лист 3: COMMERCIAL INVOICE + PACKING LIST ---
ws_ci = wb.create_sheet("COMMERCIAL INVOICE")
ws_ci.column_dimensions["A"].width = 69.8
for col in ["B","C","D","E"]:
    ws_ci.column_dimensions[col].width = 10
    ws_ci.column_dimensions["F"].width = 6
    ws_ci.column_dimensions["G"].width = 50
    for col in ["H","I","J","K","L"]:
        ws_ci.column_dimensions[col].width = 10

    # --- Заголовок CI ---
    today_str = datetime.now().strftime("%d-%m-%Y")
    ws_ci["A1"] = "COMMERCIAL INVOICE"
    ws_ci["A2"] = "Seller"
    ws_ci["A3"] = "BEELIFECOS"
    ws_ci["A4"] = "23-10 Yeongju 1(il)-dong, Jung-gu,"
    ws_ci["A5"] = "Busan, Korea"
    ws_ci["A7"], ws_ci["B7"] = "Consignee", client
    ws_ci["A8"] = client
    ws_ci["A9"] = data["address"]
    ws_ci["A13"] = "Vessel From INCHON, KOREA"
    ws_ci["B3"] = f"KR {today_str}"
    ws_ci["C3"] = ", ".join(data["order_ids"])
    ws_ci["B4"] = "L/C No. and date"
    ws_ci["B7"] = "Buyer (if other than consignee)"
    ws_ci["B8"] = "SAME AS ABOVE"
    ws_ci["B10"] = "Other references"
    ws_ci["B11"] = "ORIGIN OF KOREA"
    ws_ci["B12"] = "Terms of delivery and payment"
    ws_ci["B13"] = data["country"]

for row in ws_ci["A2:E13"]:  # диапазон можно подстроить
    for cell in row:
        cell.border = thin_border


    # Заголовки CI
    table_headers = ["Goods description", "Quantity or net weight", "EA", "Unit price", "Amount"]
    for col, val in zip(["A","B","C","D","E"], table_headers):
        ws_ci[f"{col}14"] = val
        ws_ci[f"{col}14"].font = white_font
        ws_ci[f"{col}14"].fill = header_fill
        ws_ci[f"{col}14"].alignment = center_align
        ws_ci[f"{col}14"].border = thin_border

# Данные CI
ci_start_row = 15
row_idx = ci_start_row  # вручную считаем строку

for article, name, qty_collected, qty_order, price, discount, img_url in data["rows"]:
    if qty_collected > 0:
        ws_ci[f"A{row_idx}"] = name
        ws_ci[f"B{row_idx}"] = qty_collected
        ws_ci[f"C{row_idx}"] = "EA"
        ws_ci[f"D{row_idx}"] = price-discount
        ws_ci[f"E{row_idx}"] = f"=B{row_idx}*D{row_idx}"

        for col in ["A", "B", "C", "D", "E"]:
            ws_ci[f"{col}{row_idx}"].alignment = Alignment(
                horizontal="left" if col == "A" else "center",
                vertical="center",
                wrap_text=True
            )
            ws_ci[f"{col}{row_idx}"].border = thin_border

        row_idx += 1  # увеличиваем только если строка реально записана

# Итого
last_ci_row = row_idx - 1
ws_ci[f"B{last_ci_row+1}"] = f"=SUM(B{ci_start_row}:B{last_ci_row})"
ws_ci[f"E{last_ci_row+1}"] = f"=SUM(E{ci_start_row}:E{last_ci_row})"

# Подпись
ws_ci[f"A{last_ci_row+3}"] = "Signed by"
ws_ci[f"A{last_ci_row+3}"].font = bold_font
ws_ci[f"A{last_ci_row+3}"].border = thin_border

    # --- Заголовки Packing List на том же листе ---
ws_ci["G1"] = "PACKING LIST"
ws_ci["G2"] = "Seller"
ws_ci["G3"] = "BEELIFECOS"
ws_ci["G4"] = "23-10 Yeongju 1(il)-dong, Jung-gu,"
ws_ci["G5"] = "Busan, Korea"
ws_ci["G7"] = "Consignee"
ws_ci["G8"] = client
ws_ci["G9"] = data["address"]
ws_ci["G12"] = "Departure date"
ws_ci["G13"] = "Vessel From INCHON, KOREA"
ws_ci["H3"] = f"KR {today_str}"
ws_ci["I3"] = ", ".join(data["order_ids"])
ws_ci["H4"] = "L/C No. and date"
ws_ci["H7"] = "Buyer (if other than consignee)"
ws_ci["H8"] = "SAME AS ABOVE"
ws_ci["H10"] = "Other references"
ws_ci["H11"] = "ORIGIN OF KOREA"
ws_ci["H12"] = "Terms of delivery and payment"
ws_ci["H13"] = data["country"]

# --- Обводим блок PACKING LIST (шапка) ---
for row in ws_ci["G2:L13"]:  # диапазон под себя
    for cell in row:
        cell.border = thin_border


    # Заголовки Packing List
    ws_ci["G14"] = "No. & of pkgs"
    ws_ci["H14"] = "Quantity or net weight"
    ws_ci["I14"] = "Box"
    ws_ci["J14"] = "Gross weight"
    ws_ci["K14"] = "kg"
    ws_ci["L14"] = "Mesure"
    for col in ["G","H","I","J","K","L"]:
        ws_ci[f"{col}14"].font = white_font
        ws_ci[f"{col}14"].fill = header_fill
        ws_ci[f"{col}14"].alignment = center_align
        ws_ci[f"{col}14"].border = thin_border

    # --- Данные Packing List ---
    pl_row = 15
    total_boxes, total_weight = 0, 0
    for row in data["rows"]:
        name, qty_collected = row[1], row[2]
        if qty_collected > 0:
            ws_ci[f"G{pl_row}"] = "Cosmetics"
            ws_ci[f"H{pl_row}"] = qty_collected
            ws_ci[f"I{pl_row}"] = "box"
            ws_ci[f"J{pl_row}"] = 5   # вес коробки
            ws_ci[f"K{pl_row}"] = "kg"
            ws_ci[f"L{pl_row}"] = "-"
            total_boxes += 1
            total_weight += 5
            for col in ["G","H","I","J","K","L"]:
                ws_ci[f"{col}{pl_row}"].alignment = center_align
                ws_ci[f"{col}{pl_row}"].border = thin_border
            pl_row += 1

    # --- Итоги Packing List ---
    ws_ci[f"G{pl_row}"] = "TOTAL :"
    ws_ci[f"H{pl_row}"] = f"=SUM(H15:H{pl_row-1})"
    ws_ci[f"I{pl_row}"] = "box"
    ws_ci[f"J{pl_row}"] = total_weight
    ws_ci[f"K{pl_row}"] = "kg"
    ws_ci[f"L{pl_row}"] = total_boxes
    for col in ["G","H","I","J","K","L"]:
        ws_ci[f"{col}{pl_row}"].font = bold_font
        ws_ci[f"{col}{pl_row}"].alignment = center_align
        ws_ci[f"{col}{pl_row}"].border = thin_border

# --- Загружаем данные склада ---
inventory_path = "https://docs.google.com/spreadsheets/d/e/2PACX-1vR7hj61tyAykZN2GpBmKVpvQjlAq-UJCAvc9XwyHPYTgGCdVrrTRWD5Ce-uYIZMpy4guvu6_TWjl4ca/pub?output=csv"
inventory = pd.read_csv(inventory_path)

# Переименовываем
inventory.rename(columns={
    'Box price ($)': 'box_price',
    'price ($)': 'price',
    'in box': 'in_box'
}, inplace=True)

# Проверка обязательных колонок
inventory_cols = ['code', 'barcode', 'name', 'in stock', 'category']
for col in inventory_cols:
    if col not in inventory.columns:
        raise ValueError(f"В файле отсутствует колонка: {col}")

# Приводим к числу
inventory['in stock'] = pd.to_numeric(inventory['in stock'], errors='coerce').fillna(0).astype(int)

# Добавляем недостающие колонки
for col in ['price', 'brand', 'in_box', 'box_price', 'moq']:
    if col not in inventory.columns:
        inventory[col] = '-'

# Создаем лист для закупочного инвойса
ws = wb.create_sheet("Purchase Invoice")
excel_headers = ["Order ID", "Баркод", "Название", "Артикул", "Количество", 
                 "Фактический остаток", "Нужно заказать", "Статус", "Склад",
                 "Box price ($)", "price ($)", "moq", "brand", "in box"]
ws.append(excel_headers)

# Шрифт и ширина колонок
for col_num, _ in enumerate(excel_headers, start=1):
    ws.cell(row=1, column=col_num).font = bold_font
    ws.column_dimensions[chr(64 + col_num)].width = 25

# Заливки для статусов
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Функции
def status_check(row):
    if row['in stock'] == 0:
        return "Нет в наличии"
    elif row['in stock'] >= row['Количество']:
        return "В наличии"
    else:
        return f"Частично ({int(row['in stock'])}/{int(row['Количество'])})"

def get_fill_by_status(status):
    if "Нет" in status:
        return red_fill
    elif "Частично" in status:
        return yellow_fill
    else:
        return green_fill

# Обрабатываем заказы
for oid in ORDER_IDS:
    ORDER_URL = BASE_URL + f"?dispatch=orders.print_invoice&order_id={oid}"
    response = session.get(ORDER_URL, headers=headers)
    soup = BeautifulSoup(response.text, "html.parser")
    table = soup.find("table", width="100%")
    if not table:
        continue

    rows_data = []
    for tr in table.find_all("tr")[1:]:
        tds = tr.find_all("td")
        if len(tds) < 7:
            continue
        barcode = tds[1].get_text(strip=True)
        p_tag = tds[2].find("p")
        name = p_tag.get_text(strip=True) if p_tag else tds[2].get_text(" ", strip=True)
        try:
            qty_order = int(tds[4].get_text(strip=True))
        except:
            qty_order = 0
        rows_data.append([barcode, name, qty_order])

    order_df = pd.DataFrame(rows_data, columns=["Баркод", "Название", "Количество"])

    # Объединяем с данными склада
    df = pd.merge(
        order_df,
        inventory[['barcode', 'code', 'in stock', 'category', 'price', 'brand', 'in_box', 'box_price', 'moq']],
        left_on='Баркод', right_on='barcode', how='left'
    )

    # Обработка артикулов
    df['code'] = df['code'].apply(lambda x: x.split("PartNum:")[0].strip() if pd.notna(x) and "PartNum:" in x else (x if pd.notna(x) else "-"))

    df['Количество'] = pd.to_numeric(df['Количество'], errors='coerce').fillna(0).astype(int)
    df['in stock'] = df['in stock'].fillna(0).astype(int)
    df['category'] = df['category'].fillna('-')

    # --- Корректная обработка колонок для Excel ---
    for col in ['box_price', 'price', 'moq']:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    df['Box price ($)'] = df['box_price'].apply(lambda x: x if pd.notna(x) else '-')
    df['price ($)'] = df['price'].apply(lambda x: x if pd.notna(x) else '-')
    df['moq'] = df['moq'].apply(lambda x: int(x) if pd.notna(x) else '-')
    df['brand'] = df['brand'].fillna('-')
    df['in box'] = df['in_box'].fillna('-')

    # Нужно заказать и статус
    df['Нужно заказать'] = (df['Количество'] - df['in stock']).clip(lower=0)
    df['Статус'] = df.apply(status_check, axis=1)

    # Вставка в Excel
    for _, row in df.iterrows():
        ws.append([
            oid, row['Баркод'], row['Название'], row['code'], row['Количество'],
            row['in stock'], row['Нужно заказать'], row['Статус'], row['category'],
            row['Box price ($)'], row['price ($)'], row['moq'], row['brand'], row['in box']
        ])
        current_row = ws.max_row
        fill = get_fill_by_status(row['Статус'])
        for col in range(1, len(excel_headers)+1):
            ws.cell(row=current_row, column=col).fill = fill
# --- Сохраняем файл ---
# --- Сохраняем файл локально ---
wb.save(file_path)
print(f"Файл успешно создан: {file_path}")

# --- Загружаем файл на Google Drive ---
# pip install PyDrive
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive

# 1. Авторизация
gauth = GoogleAuth()
gauth.LocalWebserverAuth()  
drive = GoogleDrive(gauth)

# 2. Папка назначения
folder_id = "1Cy1bTd2M9rdQRPjcQN4I59sPGeNfU6vh"  # ID папки
file_name = os.path.basename(file_path)

# 3. Проверка — есть ли уже файл с таким названием в папке
query = f"'{folder_id}' in parents and trashed=false and title='{file_name}'"
file_list = drive.ListFile({'q': query}).GetList()

if file_list:
    # Если нашли файл с таким названием — удаляем
    print(f"Файл '{file_name}' уже существует. Удаляю старый...")
    file_list[0].Delete()

# 4. Создаём новый файл и загружаем
file_drive = drive.CreateFile({
    'title': file_name,
    'parents': [{'id': folder_id}]
})
file_drive.SetContentFile(file_path)
file_drive.Upload()

print(f"Файл '{file_name}' успешно обновлён на Google Drive в папке {folder_id}")