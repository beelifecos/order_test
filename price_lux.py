import pandas as pd
import asyncio
import aiohttp
from io import BytesIO
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import math
from datetime import datetime
from bs4 import BeautifulSoup
import os

# PyDrive
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive

# === 0. Папка для сохранения локально ===
output_dir = os.path.join(os.path.dirname(__file__), "output")
os.makedirs(output_dir, exist_ok=True)

# === 1. Загружаем CSV ===
csv_url = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTDlbmTowkavqqkFPWzBvoT-V0eG3BFc6yy5TagW6tNC6LqpNiBzIxk9UYd1ULUfTVeTl-Yh9JXJ6Z4/pub?output=csv"
df = pd.read_csv(csv_url)
df = df.sort_values(by="brand", ascending=True)

# HTML -> текст
def html_to_text(html_content):
    soup = BeautifulSoup(str(html_content), "html.parser")
    lines = []
    for p in soup.find_all("p"):
        text = p.get_text(separator=" ").strip()
        if text:
            lines.append(text)
    return "\n".join(lines)

if 'description' in df.columns:
    df['description'] = df['description'].apply(html_to_text)

# Добавляем колонки
df['order qty'] = 0
df['total'] = 0

# === 2. Асинхронная загрузка картинок ===
async def fetch_image(session, url, retries=2):
    for attempt in range(retries):
        try:
            async with session.get(url, timeout=10) as resp:
                if resp.status == 200 and "image" in resp.headers.get("Content-Type", ""):
                    content = await resp.read()
                    img = resize_image(content)
                    if img:
                        return img
        except Exception:
            pass
    return None

def resize_image(content, width=120, height=90):
    try:
        img = PILImage.open(BytesIO(content))
        img.thumbnail((width, height))
        output = BytesIO()
        img.save(output, format="PNG")
        output.seek(0)
        return output
    except Exception:
        return None

async def download_all_images(urls, batch_size=50):
    images = {}
    urls = [url for url in urls if isinstance(url, str) and url.startswith("http")]
    for i in range(0, len(urls), batch_size):
        batch = urls[i:i+batch_size]
        async with aiohttp.ClientSession() as session:
            tasks = {url: asyncio.create_task(fetch_image(session, url)) for url in batch}
            for url, task in tasks.items():
                images[url] = await task
    return images

img_urls = df['img'].dropna().unique().tolist()
images_cache = asyncio.run(download_all_images(img_urls))

# === 3. Функция создания Excel-файла ===
def create_excel(df_part, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Price List"

    # Стили
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    highlight_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    price_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Бренды
    brands_str = ", ".join(df_part['brand'].unique())
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df_part.columns))
    brand_cell = ws.cell(row=2, column=1, value=f"Бренды: {brands_str}")
    brand_cell.font = Font(bold=True, color="000000")
    brand_cell.alignment = Alignment(horizontal='left')

    # Заголовки
    for col_num, column_title in enumerate(df_part.columns, 1):
        cell = ws.cell(row=3, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.border = border

    # Колонки для формул
    price_col = df_part.columns.get_loc('price ($)') + 1
    order_col = df_part.columns.get_loc('order qty') + 1
    total_col = df_part.columns.get_loc('total') + 1

    # Данные
    for row_num, row in enumerate(df_part.itertuples(index=False), 4):
        ws.row_dimensions[row_num].height = 90
        for col_num, value in enumerate(row, 1):
            col_name = df_part.columns[col_num-1].lower()
            cell_ref = f"{get_column_letter(col_num)}{row_num}"
            if isinstance(value, str) and value.startswith("http") and col_name == "img":
                if value in images_cache and images_cache[value]:
                    # Создаём новый поток для каждой вставки
                    img_stream = BytesIO(images_cache[value].getvalue())
                    xl_img = XLImage(img_stream)
                    xl_img.width, xl_img.height = 120, 90
                    ws.add_image(xl_img, cell_ref)
                else:
                    ws.cell(row=row_num, column=col_num, value="Image Error")
            else:
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(wrap_text=True)
                cell.border = border
                if df_part.columns[col_num-1] in ['order qty', 'total']:
                    cell.fill = highlight_fill
                if df_part.columns[col_num-1] == 'price ($)':
                    cell.fill = price_fill

    # Формулы total
    for row_num in range(4, len(df_part)+4):
        ws.cell(row=row_num, column=total_col,
                value=f"={get_column_letter(price_col)}{row_num}*{get_column_letter(order_col)}{row_num}")

    # Сумма внизу
    summary_row = len(df_part) + 4
    sum_formula = f"=SUM({get_column_letter(total_col)}4:{get_column_letter(total_col)}{summary_row-1})"
    cell = ws.cell(row=summary_row, column=total_col, value=sum_formula)
    cell.font = Font(bold=True, color="FF0000")
    cell.fill = highlight_fill
    cell.alignment = Alignment(horizontal='center')

    # Заморозка строк
    ws.freeze_panes = "A4"

    # Ширина колонок
    column_widths = {
        "A": 17, "B": 10, "C": 10, "D": 15,
        "E": 30, "F": 9, "G": 9, "H": 9,
        "I": 9, "J": 10, "K": 10, "L": 10
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Сохраняем файл
    wb.save(filename)
    print(f"✅ Файл {filename} создан с {len(df_part)} позициями!")

# === 4. Разделяем DataFrame на части по 300 позиций ===
chunk_size = 300
num_chunks = math.ceil(len(df) / chunk_size)
today = datetime.today().strftime("%Y-%m-%d")

# === 5. Авторизация Google Drive ===
gauth = GoogleAuth()
gauth.LocalWebserverAuth()  # авторизация через браузер
drive = GoogleDrive(gauth)
folder_id = "1gzwzDrK1EMVaHwux8ZI8wjmuMhR1MXRC"  # ваша папка на Google Диске

# === 6. Генерация Excel и загрузка на Google Drive ===
for i in range(num_chunks):
    start = i*chunk_size
    end = start+chunk_size
    df_chunk = df.iloc[start:end]

    brands = "_".join(df_chunk['brand'].unique())
    if len(brands) > 50:
        brands = brands[:50] + "_etc"

    filename = os.path.join(output_dir, f"price_{brands}_part{i+1}.xlsx")
    create_excel(df_chunk, filename)

    # Проверяем, есть ли файл с таким именем на диске
    query = f"'{folder_id}' in parents and trashed=false and title='{os.path.basename(filename)}'"
    file_list = drive.ListFile({'q': query}).GetList()
    if file_list:
        file_list[0].Delete()

    gfile = drive.CreateFile({'title': os.path.basename(filename), 'parents':[{'id': folder_id}]})
    gfile.SetContentFile(filename)
    gfile.Upload()
    print(f"✅ Файл '{os.path.basename(filename)}' загружен на Google Drive")
