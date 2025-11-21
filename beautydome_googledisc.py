import time
import re
import urllib.parse
import os
import pickle
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from bs4 import BeautifulSoup
from deep_translator import GoogleTranslator

from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

# ---------------- Google Drive -----------------
SCOPES = ['https://www.googleapis.com/auth/drive.file']

def drive_service():
    creds = None
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('client_secrets.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)
    service = build('drive', 'v3', credentials=creds)
    return service

def upload_to_drive(file_path, folder_id=None):
    service = drive_service()
    file_metadata = {'name': os.path.basename(file_path)}
    if folder_id:
        file_metadata['parents'] = [folder_id]
    media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file = service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink').execute()
    print(f"Файл загружен на Google Drive. ID: {file['id']}")
    print(f"Ссылка на просмотр: {file['webViewLink']}")
    return file['webViewLink']

# ---------------- Категории товаров -----------------
def assign_category(name):
    if not name:
        return "НЕОПРЕДЕЛЕНО"
    name_lower = name.lower()
    if any(k in name_lower for k in ["선크림", "sun screen", "spf", "sun cream", "sun stick", "sun care", "선스틱"]):
        return "SUN CARE I ЗАЩИТА ОТ СОЛНЦА"
    if any(k in name_lower for k in ["미셀라", "micellar", "필링","cleanser","peeling","비누","soup","코팩","nose pack","클렌징","리무버","remover","cleansing","필 오프 팩","peel off pack","폼","foam","필링 젤","peeling gel","클렌저 오일","cleanser oil","오일 클렌저","oil cleanser","마일드","mild","워터","cleansing water","water wash"]):
        return "CLEANSING I ОЧИЩЕНИЕ"
    if any(k in name_lower for k in ["앰플","ampoule","스킨","유연액","에멀션","유연수","patch","pad","reedle shot","pack","source","moisturizer","ampule","멀티밤","multi balm","에멀전","소프너","softner","크림","cream","토너","아이패치","eye patch","멀티 밤","toner","에멀젼","emulsion","엑스트라 액터","수액","리프샷","유액","마스크","mask","에센스","essence","옴므 올인원","세럼","serum","아이크림","eye cream","eye serum","하이드레이팅","hydrating","비타","vitamin","리프팅","lifting","미백","whitening","brightening","수딩","soothing","balm","concentrate","패드","링클 진액고","수분팩","앰풀오일","진액 오일","아이리프트","멀티스틱","밸런서"]):
        return "SKIN CARE I УХОД ЗА ЛИЦОМ"
    if any(k in name_lower for k in ["바디","body","로션","lotion","여성 청결제","스크럽","scrub","바디워시","넥","body wash","샤워젤","여성청결제"]):
        return "BODY CARE I УХОД ЗА ТЕЛОМ"
    if any(k in name_lower for k in ["샴푸","shampoo","컨디셔너","conditioner","헤어","hair","트리트먼트","treatment","헤어팩","hair pack","헤어오일","hair oil"]):
        return "HAIR CARE I УХОД ЗА ВОЛОСАМИ"
    if any(k in name_lower for k in ["립","lip","파운데이션","foundation","블러셔","blush","섀도우","eyeshadow","마스카라","mascara","bb cream","아이브로우","eye brow","팩트","pact","파우더","powder","틴트","tint"]):
        return "MAKE UP I ДЕКОРАТИВНЫЙ МАКИЯЖ"
    if any(k in name_lower for k in ["세트","set","special set","패키지","package","컬렉션","collection","kit","키트"]):
        return "SKIN CARE SET I УХОДОВЫЕ НАБОРЫ"
    if any(k in name_lower for k in ["남성","men","for men","homme"]):
        return "FOR MEN / Для мужчин"
    if any(k in name_lower for k in ["샘플","sample","미니","mini","트래블","travel"]):
        return "SAMPLE | ПРОБНИКИ"
    if any(k in name_lower for k in ["건강기능식품","supplement","비타민","vitamin","오메가","omega","프로바이오틱스","probiotic"]):
        return "БАДЫ"
    if any(k in name_lower for k in ["코롱","perfume","향수","치약","손소독제"]):
        return "ТОВАРЫ ДЛЯ ДОМА И ЗДОРОВЬЯ"
    return "НЕОПРЕДЕЛЕНО"

# ---------------- Бренды -----------------
brand_name_map = {
    "1703": {"ko": "가인비책", "en": "GAINBICHAEK"},
    "1700": {"ko": "더후(더히스토리오브후)", "en": "THE HISTORY OF WHOO"},
    "1250": {"ko": "과일나라", "en": "FRUIT NARA"},
    "1252": {"ko": "꽃을든남자", "en": "FLOWER MAN"},
    "1668": {"ko": "끌레드벨", "en": "CLE DE BELLE"},
    "1253": {"ko": "나드리 이노벨라", "en": "NADRI INNOVELLA"},
    "1256": {"ko": "뉴트로지나", "en": "NEUTROGENA"},
    "1257": {"ko": "다나한", "en": "DANAHAN"},
    "1671": {"ko": "도루코", "en": "DORCO"},
    "1689": {"ko": "동성제약", "en": "DONGSUNG PHARM"},
    "1498": {"ko": "드봉", "en": "DEBON"},
    "1268": {"ko": "라미", "en": "RAMI"},
    "1273": {"ko": "루나리스", "en": "LUNARIS"},
    "1424": {"ko": "리엔", "en": "RIEN"},
    "1462": {"ko": "릴랙시아", "en": "RELAXIA"},
    "1280": {"ko": "마몽드", "en": "MAMONDE"},
    "1283": {"ko": "멘소래담", "en": "MENTHOLATUM"},
    "1284": {"ko": "무궁화", "en": "MUGUNGHWA"},
    "1442": {"ko": "미쟝센", "en": "MISE EN SCENE"},
    "1289": {"ko": "바세린", "en": "VASELINE"},
    "1290": {"ko": "바찌", "en": "BAZZI"},
    "1291": {"ko": "백옥생", "en": "BAEKOKSAENG"},
    "1713": {"ko": "베르가모", "en": "BERGAMO"},
    "1736": {"ko": "브에노", "en": "BUENO"},
    "1741": {"ko": "비노아", "en": "VINNOA"},
    "1292": {"ko": "비러브", "en": "BELUV"},
    "1299": {"ko": "산수유", "en": "SANSUYU"},
    "1747": {"ko": "설려", "en": "SEOLLYO"},
    "1494": {"ko": "소망기타", "en": "SOMANG"},
    "1684": {"ko": "숨37도", "en": "SUM37"},
    "1489": {"ko": "쉬림", "en": "SHRIM"},
    "1664": {"ko": "쉬크", "en": "SCHICK"},
    "1702": {"ko": "썬월드", "en": "SUNWORLD"},
    "1734": {"ko": "씨드앤팜", "en": "SEED&PHARM"},
    "1303": {"ko": "아방가드로", "en": "AVANGARDO"},
    "1727": {"ko": "아이차밍", "en": "ICHARMING"},
    "1701": {"ko": "아트피아", "en": "ARTPIA"},
    "1706": {"ko": "알프레도 휘마스", "en": "ALFREDO HUIMAS"},
    "1496": {"ko": "애경", "en": "AEKYUNG"},
    "1676": {"ko": "에띠앙", "en": "ETTIANG"},
    "1308": {"ko": "에바스", "en": "EBAS"},
    "1309": {"ko": "에뿌", "en": "EPPU"},
    "1312": {"ko": "에스클라", "en": "ESCLA"},
    "1678": {"ko": "에스클레어", "en": "ESCLAIR"},
    "1313": {"ko": "에이쓰리에프온", "en": "A3FON"},
    "1315": {"ko": "에코퓨어", "en": "ECO PURE"},
    "1316": {"ko": "엔프라니", "en": "ENPRANI"},
    "1317": {"ko": "엘라스틴", "en": "ELASTINE"},
    "1729": {"ko": "엘지생활건강", "en": "LG HOUSEHOLD & HEALTH"},
    "1711": {"ko": "예지후", "en": "YEJIHU"},
    "1737": {"ko": "예향", "en": "YEHYANG"},
    "1732": {"ko": "오가니아", "en": "OGANIA"},
    "1318": {"ko": "오딧세이", "en": "ODYSSEY"},
    "1714": {"ko": "오릭스", "en": "ORIX"},
    "1320": {"ko": "오퍼스", "en": "OPUS"},
    "1673": {"ko": "오휘(O HUI)", "en": "O HUI"},
    "1321": {"ko": "온더바디", "en": "ON THE BODY"},
    "1322": {"ko": "우드버리", "en": "WOODBURY"},
    "1726": {"ko": "이노벨라", "en": "INNOVELLA"},
    "1440": {"ko": "존슨앤존슨", "en": "JOHNSON & JOHNSON"},
    "1421": {"ko": "쥬리아", "en": "JULIA"},
    "1327": {"ko": "지오", "en": "GIO"},
    "1712": {"ko": "카라코사", "en": "KARAKOSA"},
    "1337": {"ko": "터치미", "en": "TOUCH ME"},
    "1725": {"ko": "팜스테이(명인화장품)", "en": "FARM STAY"},
    "1748": {"ko": "포더스킨", "en": "FOR THE SKIN"},
    "1345": {"ko": "푸드어홀릭", "en": "FOODAHOLIC"},
    "1423": {"ko": "프린시아", "en": "PRINCIA"},
    "1347": {"ko": "피어리스", "en": "PEERLESS"},
    "1349": {"ko": "한불", "en": "HANBUL"},
    "1491": {"ko": "해피바스", "en": "HAPPY BATH"},
    "1495": {"ko": "황후빈", "en": "HWANGHUBIN"}
    # ... добавь остальные бренды ...
}

def extract_brand_name(brand_url):
    query = urllib.parse.urlparse(brand_url).query
    params = urllib.parse.parse_qs(query)
    brand_cd = params.get("cno1", [""])[0]
    brand_info = brand_name_map.get(brand_cd, {"ko": "Unknown Brand", "en": "Unknown Brand"})
    return brand_info["ko"], brand_info["en"]

def brand_special_column(brand_name_ko, brand_name_en):
    if not brand_name_en or brand_name_en.strip() == "":
        return f"{brand_name_ko}///X///UNKNOWN"
    first_letter = brand_name_en.strip()[0].upper()
    return f"Бренд///{first_letter}///{brand_name_en.strip().upper()}"

def handle_alert(driver):
    try:
        WebDriverWait(driver, 3).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        alert.accept()
    except:
        pass

def add_page_to_url(url, page_num):
    if "page=" in url:
        return re.sub(r'page=\d+', f'page={page_num}', url)
    separator = '&' if '?' in url else '?'
    return f"{url}{separator}page={page_num}"

def safe_str_price(value):
    return f"{value:.2f}".replace(",", ".") if value is not None else ""

def translate_name_to_en(korean_name):
    try:
        if not korean_name or korean_name.strip() == "":
            return ""
        translated = GoogleTranslator(source='ko', target='en').translate(korean_name)
        return translated.strip().capitalize()
    except Exception as e:
        print(f"Ошибка перевода '{korean_name}': {e}")
        return ""

# ---------------- Scraper -----------------
def login_and_scrape(username, password):
    options = Options()
    options.add_argument('--disable-notifications')
    options.add_argument("start-maximized")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    # Вход
    driver.get("https://www.beautydome.co.kr/member/login.php")
    handle_alert(driver)
    driver.find_element(By.ID, "login_id").send_keys(username)
    driver.find_element(By.ID, "login_pwd").send_keys(password)
    driver.find_element(By.CSS_SELECTOR, ".box_btn.circle input[type='submit']").click()
    handle_alert(driver)

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Изображение", "Бренд", "Название", "name_en", "Единица измерения", "MOQ", "Остаток",
        "Цена", "Цена розницы", "Артикул", "Excel формула", "Категория",
        "Language", "Lower limit", "User group", "Особенности","brand_name_en", "price","cena_na_site","Status",
    ])

    seen_items = set()

    brand_urls = [f"https://www.beautydome.co.kr/shop/big_section.php?cno1={key}" for key in brand_name_map.keys()]

    for brand_url in brand_urls:
        brand_name_ko, brand_name_en = extract_brand_name(brand_url)
        brand_column = brand_special_column(brand_name_ko, brand_name_en)
        print(f"Scraping products for brand: {brand_column}")

        for page_num in range(1, 11):
            page_url = add_page_to_url(brand_url, page_num)
            driver.get(page_url)
            time.sleep(3)

            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.info"))
                )
            except:
                print(f"Страница {page_num} бренда {brand_column} пустая. Прерываем цикл.")
                break

            soup = BeautifulSoup(driver.page_source, 'html.parser')
            products = soup.select("div.info")
            if not products:
                print(f"Страница {page_num} бренда {brand_column} пустая. Прерываем цикл.")
                break

            print(f"Страница {page_num} бренда {brand_column} - найдено {len(products)} товаров")

            for product in products:
                try:
                    name_tag = product.select_one("p.name a")
                    if not name_tag:
                        continue
                    href = name_tag['href']
                    params = urllib.parse.parse_qs(urllib.parse.urlparse(href).query)
                    item_code = params.get('pno', [''])[0]
                    if item_code in seen_items:
                        continue
                    seen_items.add(item_code)

                    img_tag = product.select_one("div.img a img")
                    img_src = img_tag['src'] if img_tag else None
                    name = name_tag.get_text(strip=True)

                    price_old_tag = product.select_one("ul.prc .normal_prc")
                    price_discounted_tag = product.select_one("ul.prc strong")
                    price_old = price_old_tag.get_text(strip=True).replace(",", "").replace("원", "") if price_old_tag else None
                    price_discounted = price_discounted_tag.get_text(strip=True).replace(",", "").replace("원", "") if price_discounted_tag else None

                    if price_discounted:
                        price_discounted_int = int(re.sub(r"[^\d]", "", price_discounted))
                        price = round(price_discounted_int * 1.15 / 1250, 2)
                        cena_na_site = round(price_discounted_int * 1.3 / 1250, 2)
                    else:
                        price = None
                        cena_na_site = None

                    price_str = safe_str_price(price)
                    cena_na_site_str = safe_str_price(cena_na_site)
                    name_en = translate_name_to_en(name)

                    moq = None
                    quantity_avail = "20"
                    category = assign_category(name)

                    ws.append([
                        img_src, brand_column, name, name_en, 'ea', moq, quantity_avail,
                        price_discounted, price_old, item_code,
                        f'=R{ws.max_row}&"/"&B{ws.max_row}', category, "ru",
                        20, "Все", 3, brand_name_en, price_str, cena_na_site_str, "A",
                    ])
                except Exception as e:
                    print("Ошибка при разборе товара:", e)

    driver.quit()

    file_path = f"beautydome_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(file_path)
    print(f"Файл локально сохранен: {file_path}")
    link = upload_to_drive(file_path, folder_id="10J-E4RcBJFfrdcqU_UAWask8BKTZ5Mw2")
    print(f"Ссылка на Google Drive: {link}")

# ---------------- Запуск -----------------
login_and_scrape("beelifecos", "lapulik1983*")
