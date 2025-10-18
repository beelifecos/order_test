import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import os
import subprocess
import sys

# --- Настройки ---
BASE_URL = "https://www.beelifecos.com/beelifecos_admin.php"
USERNAME = "beelifecos@gmail.com"
PASSWORD = "pompusik1986*"

# Список заказов для обработки
ORDER_IDS = [3843]  # добавь свои order_id

# CSV из Google Sheets
inventory_path = "https://docs.google.com/spreadsheets/d/e/2PACX-1vR7hj61tyAykZN2GpBmKVpvQjlAq-UJCAvc9XwyHPYTgGCdVrrTRWD5Ce-uYIZMpy4guvu6_TWjl4ca/pub?output=csv"

# --- Загружаем данные склада ---
inventory = pd.read_csv(inventory_path)

# Проверка обязательных колонок
inventory_cols = ['code', 'barcode', 'name', 'in stock', 'category']
for col in inventory_cols:
    if col not in inventory.columns:
        raise ValueError(f"В файле отсутствует колонка: {col}")

# --- Преобразуем числовые колонки ---
inventory['in stock'] = pd.to_numeric(inventory['in stock'], errors='coerce').fillna(0).astype(int)

# --- Настройка сессии ---
session = requests.Session()
headers = {"User-Agent": "Mozilla/5.0"}

# Авторизация
resp = session.get(BASE_URL, headers=headers)
soup = BeautifulSoup(resp.text, "html.parser")
form = soup.find("form", {"name": "main_login_form"})
login_data = {inp.get("name"): inp.get("value", "") for inp in form.find_all("input") if inp.get("name")}
login_data["user_login"] = USERNAME
login_data["password"] = PASSWORD
login_data["dispatch[auth.login]"] = "Войти"
session.post(BASE_URL, data=login_data, headers=headers)

# --- Создаем Excel ---
downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
report_path = os.path.join(downloads_path, "invoice_all_orders.xlsx")
wb = Workbook()
ws = wb.active
ws.title = "Инвойс"

headers_excel = ["Order ID", "Баркод", "Название", "Артикул", "Количество", 
                 "Фактический остаток", "Нужно заказать", "Статус", "Склад"]
ws.append(headers_excel)

# --- Стили ---
bold_font = Font(bold=True)
for col in range(1, len(headers_excel)+1):
    ws.cell(row=1, column=col).font = bold_font
    ws.column_dimensions[chr(64+col)].width = 25

red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# --- Функция статуса наличия ---
def status_check(row):
    if row['in stock'] == 0:
        return "Нет в наличии"
    elif row['in stock'] >= row['Количество']:
        return "В наличии"
    else:
        return f"Частично ({int(row['in stock'])}/{int(row['Количество'])})"

# --- Парсим каждый заказ ---
for ORDER_ID in ORDER_IDS:
    ORDER_URL = BASE_URL + f"?dispatch=orders.print_invoice&order_id={ORDER_ID}"
    response = session.get(ORDER_URL, headers=headers)
    soup = BeautifulSoup(response.text, "html.parser")
    table = soup.find("table", width="100%")
    if not table:
        print(f"⚠ Не удалось найти таблицу заказа {ORDER_ID}. Пропускаем.")
        continue

    rows_data = []
    for tr in table.find_all("tr")[1:]:
        tds = tr.find_all("td")
        if len(tds) < 7:
            continue
        barcode = tds[1].get_text(strip=True)
        p_tag = tds[2].find("p")
        name = p_tag.get_text(strip=True) if p_tag else tds[2].get_text(" ", strip=True)
        try:
            qty_order = int(tds[4].get_text(strip=True))
        except ValueError:
            qty_order = 0
        rows_data.append([barcode, name, qty_order])

    order_df = pd.DataFrame(rows_data, columns=["Баркод", "Название", "Количество"])

    # --- Объединяем заказ со складом ---
    df = pd.merge(order_df, inventory[['barcode', 'code', 'in stock', 'category']], 
                  left_on='Баркод', right_on='barcode', how='left')

    # --- Корректируем значения ---
# Оставляем только часть до "PartNum:"
    def extract_partnum(code):
     if pd.isna(code):
        return "-"
     return code.split("PartNum:")[0].strip() if "PartNum:" in code else code

    df['code'] = df['code'].apply(extract_partnum)
    df['Количество'] = pd.to_numeric(df['Количество'], errors='coerce').fillna(0).astype(int)
    df['in stock'] = df['in stock'].fillna(0)
    df['category'] = df['category'].fillna('-')
    # --- Рассчет нужного количества ---
    df['Нужно заказать'] = df['Количество'] - df['in stock']
    df['Нужно заказать'] = df['Нужно заказать'].apply(lambda x: x if x > 0 else 0)

    # --- Статус наличия ---
    df['Статус'] = df.apply(status_check, axis=1)

    # --- Заполняем Excel ---
    for idx, row in df.iterrows():
        ws.append([ORDER_ID, row['Баркод'], row['Название'], row['code'], row['Количество'],
                   row['in stock'], row['Нужно заказать'], row['Статус'], row['category']])
        current_row = ws.max_row
        status = row['Статус']
        fill = None
        if status == "Нет в наличии":
            fill = red_fill
        elif "Частично" in str(status):
            fill = yellow_fill
        elif status == "В наличии":
            fill = green_fill
        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=current_row, column=col).fill = fill

# --- Сохраняем и открываем Excel ---
wb.save(report_path)

if sys.platform == "darwin":  # macOS
    subprocess.call(["open", report_path])
elif sys.platform == "win32":  # Windows
    os.startfile(report_path)
elif sys.platform.startswith("linux"):  # Linux
    subprocess.call(["xdg-open", report_path])

print(f"✅ Инвойс для всех заказов создан и открыт: {report_path}")
